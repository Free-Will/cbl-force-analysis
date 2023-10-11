# Code tested on python 3.10, ensure you have that version installed on your machine before running

# If running for the first time, run the following commands in your terminal:
# pip install opencv-python
# pip install opencv-python-headless
# pip install tk
# pip install openpyxl
import os
import cv2
import openpyxl
from tkinter import filedialog as fd

#############
### This first section of code just deals with reading in the video file, parsing it to images, and preparing the directory
#############

# Create a file dialog to ask for the video file to analyze
file_dialog = fd.askopenfilename()

# Open the video file
video = cv2.VideoCapture(file_dialog)

### Chop up the entire file path to get the file name
# Split the filepath at the last "/" character
split_filepath = file_dialog.split("/")

# Get the last element of the split filepath, which is the filename
filename = split_filepath[-1]

# Split the filename at the "." character
split_filename = filename.split(".")

# Get the first element of the split filename, which is the filename without the extension
filename_without_extension = split_filename[0]

# Create a directory to store the output images
# first use of f strings, google "python f strings" if unfamiliar
output_dir = f"{filename_without_extension}_frames"
# Only creates the directory if not there.
if not os.path.exists(output_dir):
    os.makedirs(output_dir)

### Iterate over the frames of the video and save them as images
frame_total = 0
frame_file_names = []
while True:
    # Capture the next frame
    # video.read() returns a boolean value, and the next frame of the video
    # The boolean value returns False if there are no more frames to read
    ret, frame = video.read()

    # If the end of the video is reached, break the while loop
    if not ret:
        break

    # Increment our frame_total counter
    frame_total += 1
    # Save the frame as an image
    output_filename = f"{output_dir}/orig_frame_{video.get(cv2.CAP_PROP_POS_FRAMES)}.jpg"
    cv2.imwrite(output_filename, frame)
    # Save the frame file name to our list
    frame_file_names.append(output_filename)

# Release the video capture object
video.release()

#############
### This section of code handles the image analysis, frame by frame, and saves the data
#############

#TODO build input to Ask / confirm all parameters
hough_dp_param = 1
hough_minDist_param = 50
hough_param1 = 15
hough_param2 = 20
hough_minRadius = 25
hough_maxRadius = 35

microns_per_pixel = 0.23214
spring_constant = 29.49

posts_coords = {}
posts_deflection = {}
posts_force = {}
posts_orig = []
frame_count = 0
for current_frame_name in frame_file_names:
    frame_count += 1
    # Read in the image
    original_image = cv2.imread(current_frame_name)
    # Convert the image to grayscale, the Hough transform only accepts grayscale
    gray_image = cv2.cvtColor(original_image, cv2.COLOR_BGR2GRAY)

    # Apply the Hough transform to detect circles
    # Full documentation here: https://docs.opencv.org/4.x/dd/d1a/group__imgproc__feature.html#ga47849c3be0d0406ad3ca45db65a25d2d
    # Recommended to look at the full documentation to understand the input parameters
    current_circles = cv2.HoughCircles(gray_image, cv2.HOUGH_GRADIENT, dp=hough_dp_param, minDist=hough_minDist_param,
                                       param1=hough_param1, param2=hough_param2, minRadius=hough_minRadius, maxRadius=hough_maxRadius)

    # Draw the detected circles on the image
    all_centers = []
    circled_image = original_image
    if current_circles is not None:
        for current_circle in current_circles[0]:
            # Convert the the circle data to integers
            intX, intY, intR = int(current_circle[0]), int(current_circle[1]), int(current_circle[2])
            current_center = [intX,intY]
            all_centers.append(current_center)
            cv2.circle(circled_image, (intX, intY), intR, (0, 255, 0), 2)

    # Sort all of the center points. This only sorts them by the x value
    all_centers = sorted(all_centers)
    # Create an empty list for the properly sorted values
    all_centers_sorted = []

    ## This section gets all of the circle data, sorts it,  and does some math to store deflection and force

    # If we're on the first frame, we need to do some things manually
    if frame_count == 1:

        # We know the first 3 values are in the first column, thanks to sorting by X 
        # We place them into a new list 'first_col', once we sort those 3 values by their Y value
        first_col = sorted((all_centers[0],all_centers[1],all_centers[2]), key=lambda x: x[1])
        # Then add the newly sorted values into the fully sorted list
        for center in first_col:
            all_centers_sorted.append(center)

        # Repeat for the remaining columns / values
        second_col = sorted((all_centers[3],all_centers[4],all_centers[5]), key=lambda x: x[1])
        for center in second_col:
            all_centers_sorted.append(center)

        third_col = sorted((all_centers[6],all_centers[7],all_centers[8]), key=lambda x: x[1])
        for center in third_col:
            all_centers_sorted.append(center)
        
        fourth_col = sorted((all_centers[9],all_centers[10],all_centers[11]), key=lambda x: x[1])
        for center in fourth_col:
            all_centers_sorted.append(center)

        # We need to populate the posts dictionary with these newly sorted values
        # We know the 'all_centers_sorted' list is now in the same order as our posts
        # Iterate through them to create our initial posts_coords dictionary and posts_orig values in the correct order
        postCounter = 1
        for center in all_centers_sorted:
            postName = f'post{postCounter}'
            posts_coords[postName] = [center]
            posts_orig.append(center)
            posts_deflection[postName] = [[posts_coords.get(postName)[0][0]-center[0],posts_coords[postName][0][1]-center[1]]]
            posts_force[postName] = [[posts_deflection[postName][0][0] * microns_per_pixel * spring_constant, posts_deflection[postName][0][1] * microns_per_pixel * spring_constant]]
            postCounter += 1

    # For every other frame besides the first one, we need to dynamically update our post data based on 
    # the unsorted, and potentially missing center data
    else:
        
        # Iterate through all of the centers we have, this may be less than 12
        for center in all_centers:
            centerX = center[0]
            centerY = center[1]

            # For each new X,Y coord, now we'll iterate through the original X,Y coords of the posts
            for post_coord_orig in posts_orig:
                x_check = False
                y_check = False
                # Check if this X value is within 10 pixels of an original post's X value
                if abs(centerX - post_coord_orig[0]) < 25:
                    x_check = True
                # Check if this Y value is within 10 pixels of an original post's Y Value
                if abs(centerY - post_coord_orig[1]) < 25:
                    y_check = True
                
                # If this center is within 10 pixels of both the X and Y values of a given post's original location, we can 
                # safely say they are the same post. We will update that posts data with this center data
                if x_check and y_check:
                    postNumber = posts_orig.index(post_coord_orig) + 1
                    postName = f'post{postNumber}'
                    posts_coords[postName].append(center)
                    # Store the deflection value for this frame into the posts_deflection dictionary
                    posts_deflection[postName].append([centerX-post_coord_orig[0], centerY-post_coord_orig[1]])
                    break
        
        # If we didn't have 12 centers, some of our posts will be missing data
        # Find the posts that don't have data for every frame, and add an "NA"
        for post in posts_coords:
            if len(posts_coords[post]) < frame_count:
                posts_coords[post].append("NA")
            if len(posts_deflection[post]) < frame_count:
                posts_deflection[post].append("NA")
            
        # Populate the posts_force dictionary with the force calculations of each frame
        for post in posts_coords:
            try:
                current_index = frame_count - 1
                posts_force[post].append([posts_deflection.get(post)[current_index][0] * microns_per_pixel * spring_constant, posts_deflection.get(post)[current_index][1] * microns_per_pixel * spring_constant])
            except TypeError:
                posts_force[post].append('NA')

    # Save the new circled image
    file_name_no_jpg = current_frame_name.split(".")[0]
    circled_filename = f"{file_name_no_jpg}_circles.jpg"
    cv2.imwrite(circled_filename, circled_image)

#########
### Dump post coordinates into excel
#########

wb = openpyxl.Workbook()
center_sheet = wb.active
center_sheet.title = "Center Points"
deflection_sheet = wb.create_sheet("Deflection Data")
force_sheet = wb.create_sheet("Force Calculation")

### Populate "Center Points" sheet with appropriate data
# Write the "Frame 1" text down the Y axis
for frame in range(frame_count):
    center_sheet.cell(row=frame+2,column=1).value = f"Frame {frame+1}"

# Write the "Post 1 - X" text across the X axis
post_number = 1
for post in posts_coords:
    for x in range(2):
        if x == 0:
            col_temp = post_number+1
            center_sheet.cell(row=1,column=col_temp).value = f"{post} - X"
        if x == 1:
            col_temp = post_number+2
            center_sheet.cell(row=1,column=col_temp).value = f"{post} - Y"
    post_number += 2

# Actually populate the data for each cell
post_number = 1
for post in posts_coords:
    post_frame = 1
    for post_data in posts_coords[post]:
        row_temp = post_frame+1
        for data in post_data:
            data_index = post_data.index(data)
            if data_index == 0:
                col_temp = post_number+1
                center_sheet.cell(row=row_temp, column=col_temp).value = f"{data}"
            if data_index == 1:
                col_temp = post_number+2
                center_sheet.cell(row=row_temp, column=col_temp).value = f"{data}"
        post_frame +=1
    post_number += 2

### Populate "Deflection Data" sheet with appropriate data

# Write the "Frame 1" text down the Y axis
for frame in range(frame_count):
    deflection_sheet.cell(row=frame+2,column=1).value = f"Frame {frame+1}"

# Write the "Post 1 - X" text across the X axis
post_number = 1
for post in posts_deflection:
    for x in range(2):
        if x == 0:
            col_temp = post_number+1
            deflection_sheet.cell(row=1,column=col_temp).value = f"{post} - X"
        if x == 1:
            col_temp = post_number+2
            deflection_sheet.cell(row=1,column=col_temp).value = f"{post} - Y"
    post_number += 2

# Actually populate the data for each cell
post_number = 1
for post in posts_deflection:
    post_frame = 1
    for post_data in posts_deflection[post]:
        row_temp = post_frame+1
        data_index = 0
        for data in post_data:
            if data_index == 0:
                col_temp = post_number+1
                deflection_sheet.cell(row=row_temp, column=col_temp).value = f"{data}"
                data_index += 1
            if data_index == 1:
                col_temp = post_number+2
                deflection_sheet.cell(row=row_temp, column=col_temp).value = f"{data}"
        post_frame +=1
    post_number += 2

### Populate "Force Calculation" sheet with appropriate data

# Write the "Frame 1" text down the Y axis
for frame in range(frame_count):
    force_sheet.cell(row=frame+2,column=1).value = f"Frame {frame+1}"

# Write the "Post 1 - X" text across the X axis
post_number = 1
for post in posts_force:
    for x in range(2):
        if x == 0:
            col_temp = post_number+1
            force_sheet.cell(row=1,column=col_temp).value = f"{post} - X"
        if x == 1:
            col_temp = post_number+2
            force_sheet.cell(row=1,column=col_temp).value = f"{post} - Y"
    if post_number == 25:
        col_temp += 1
        force_sheet.cell(row=1,column=col_temp).value = "Force Total"
    post_number += 2 

# Actually populate the data for each cell
post_number = 1
for post in posts_force:
    post_frame = 1
    for post_data in posts_force[post]:
        row_temp = post_frame+1
        data_index = 0
        for data in post_data:
            if data_index == 0:
                col_temp = post_number+1
                force_sheet.cell(row=row_temp, column=col_temp).value = f"{data}"
                data_index += 1
            if data_index == 1:
                col_temp = post_number+2
                force_sheet.cell(row=row_temp, column=col_temp).value = f"{data}"
        post_frame +=1
    post_number += 2

# Save the workbook with mostly the same name as the original video file
wb.save(f'{filename_without_extension}_data.xlsx')